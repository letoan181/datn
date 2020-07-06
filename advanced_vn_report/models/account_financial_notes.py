import base64
import datetime

from datetime import datetime, timedelta, time
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from odoo import models, fields, api
from odoo.modules.module import get_module_resource


class FinancialNotes(models.Model):
    _name = 'financial.note'

    name = fields.Char("Tên", required=True)

    # general
    def _get_general_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'general')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['general'])]

    financial_config_lines_general = fields.One2many('financial.config.line', 'financial_note_id',
                                                     domain=lambda self: self._get_general_domain())

    # replenish
    def _get_replenish_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'replenish')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['replenish'])]

    financial_config_lines_replenish = fields.One2many('financial.config.line', 'financial_note_id',
                                                       domain=lambda self: self._get_replenish_domain())

    # invest
    def _get_invest_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'invest')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['invest'])]

    financial_config_lines_invest = fields.One2many('financial.config.line', 'financial_note_id',
                                                    domain=lambda self: self._get_invest_domain())

    # other
    def _get_other_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'other')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['other'])]

    financial_config_lines_other = fields.One2many('financial.config.line', 'financial_note_id',
                                                   domain=lambda self: self._get_other_domain())

    # bad_debit
    def _get_bad_debit_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'bad_debit')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['bad_debit'])]

    financial_config_lines_bad_debit = fields.One2many('financial.config.line', 'financial_note_id',
                                                       domain=lambda self: self._get_bad_debit_domain())

    # inventory
    def _get_inventory_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'inventory')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['inventory'])]

    financial_config_lines_inventory = fields.One2many('financial.config.line', 'financial_note_id',
                                                       domain=lambda self: self._get_inventory_domain())

    # long_asset
    def _get_long_asset_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'long_asset')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['long_asset'])]

    financial_config_lines_long_asset = fields.One2many('financial.config.line', 'financial_note_id',
                                                        domain=lambda self: self._get_long_asset_domain())

    # non_virtual_assets
    def _get_non_virtual_assets_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'non_virtual_assets')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['non_virtual_assets'])]

    financial_config_lines_non_virtual_assets = fields.One2many('financial.config.line', 'financial_note_id',
                                                                domain=lambda self: self._get_non_virtual_assets_domain())

    # virtual_assets
    def _get_virtual_assets_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'virtual_assets')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['virtual_assets'])]

    financial_config_lines_virtual_assets = fields.One2many('financial.config.line', 'financial_note_id',
                                                            domain=lambda self: self._get_virtual_assets_domain())

    # asset_rest
    def _get_asset_rest_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'asset_rest')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['asset_rest'])]

    financial_config_lines_asset_rest = fields.One2many('financial.config.line', 'financial_note_id',
                                                        domain=lambda self: self._get_asset_rest_domain())

    # bds_invest
    def _get_bds_invest_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'bds_invest')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['bds_invest'])]

    financial_config_lines_bds_invest = fields.One2many('financial.config.line', 'financial_note_id',
                                                        domain=lambda self: self._get_bds_invest_domain())

    # borrow
    def _get_borrow_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'borrow')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['borrow'])]

    financial_config_lines_borrow = fields.One2many('financial.config.line', 'financial_note_id',
                                                    domain=lambda self: self._get_borrow_domain())

    # rest_debit
    def _get_rest_debit_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'rest_debit')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['rest_debit'])]

    financial_config_lines_rest_debit = fields.One2many('financial.config.line', 'financial_note_id',
                                                        domain=lambda self: self._get_rest_debit_domain())

    # vat_other
    def _get_vat_other_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'vat_other')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['vat_other'])]

    financial_config_lines_vat_other = fields.One2many('financial.config.line', 'financial_note_id',
                                                       domain=lambda self: self._get_vat_other_domain())

    # trai_phieu
    def _get_trai_phieu_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'trai_phieu')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['trai_phieu'])]

    financial_config_lines_trai_phieu = fields.One2many('financial.config.line', 'financial_note_id',
                                                        domain=lambda self: self._get_trai_phieu_domain())

    # csh
    def _get_csh_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'csh')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['csh'])]

    financial_config_lines_csh = fields.One2many('financial.config.line', 'financial_note_id',
                                                 domain=lambda self: self._get_csh_domain())

    # assets_wait
    def _get_assets_wait_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'assets_wait')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['assets_wait'])]

    financial_config_lines_assets_wait = fields.One2many('financial.config.line', 'financial_note_id',
                                                         domain=lambda self: self._get_assets_wait_domain())

    # customer_pay
    def _get_customer_pay_domain(self):
        # domain = []
        # lines = self.env['financial.config.line'].sudo().search([('page', '=', 'customer_pay')])
        # for line in lines:
        #     domain.append(line.id)
        # return [('id', 'in', domain)]
        return [('page', 'in', ['customer_pay'])]

    financial_config_lines_customer_pay = fields.One2many('financial.config.line', 'financial_note_id',
                                                          domain=lambda self: self._get_customer_pay_domain())

    def save(self):
        pass

    @api.model
    def update_sequence(self):
        all_null_sequence = self.env['financial.config.line'].sudo().search([('sequence', '=', False)])
        if len(all_null_sequence) > 0:
            for line in all_null_sequence:
                line.sudo().write({
                    'sequence': line.id
                })
        return True


class FinancialNoteLines(models.Model):
    _name = 'financial.config.line'
    # _order = 'complete_code_name'
    # _order = 'code'
    _rec_name = 'code'
    _order = 'sequence,id'

    name = fields.Text('Tên chỉ tiêu')
    code = fields.Char('Mã chỉ tiêu')
    # complete_code_name = fields.Char('Mã chỉ tiêu chi tiết', compute='_compute_complete_name', store=True)
    financial_note_id = fields.Many2one('financial.note', ondelete='cascade')
    sequence = fields.Integer()

    # parent_id = fields.Many2one('financial.config.line', index=True, ondelete='cascade',string='Thuộc phần')
    parent_id = fields.Char(string='Thuộc phần', compute='onchange_parent_id')

    @api.depends('code')
    def onchange_parent_id(self):
        for rec in self:
            rec.parent_id = False
            if rec.code and '.' in rec.code:
                parent_code = rec.code.split('.')[0]
                rec.parent_id = rec.code.split('.')[0]
            else:
                rec.parent_id = False

    col = fields.Char("Thuộc cột")
    row = fields.Char("Thuộc hàng")
    type = fields.Selection([('tieu_de', 'Tiêu đề'), ('chi_tiet', 'Chi tiết'), ('tong_hop', 'Tổng hợp'),
                             ('chu_thich', 'Chú thích')
                             ], string="Loại chỉ tiêu")
    page = fields.Selection([('general', 'Genneral'),
                             ('replenish', 'Replenish'),
                             ('invest', 'Invest'),
                             ('other', 'Other'),
                             ('bad_debit', 'Bad Debit'),
                             ('inventory', 'Inventory'),
                             ('long_asset', 'Long Assets'),
                             ('non_virtual_assets', 'TSCĐ hữu hình'),
                             ('virtual_assets', 'TSCĐ vô hình'),
                             ('asset_rest', 'Tài sản thuê'),
                             ('bds_invest', 'Tài sản BĐS đâu tư'),
                             ('borrow', 'Các khoản vay'),
                             ('rest_debit', 'Nợ thuê tài chính'),
                             ('vat_other', 'Thuế và các khoản nộp'),
                             ('trai_phieu', 'Tri phieu phat hanh'),
                             ('csh', 'Biến động CSH'),
                             ('assets_wait', 'Tài sản chờ xử lý'),
                             ('customer_pay', 'Phải trả khách hàng'),
                             ], string="Trang")
    note = fields.Text('Nội dung')
    initial = fields.Many2one('financial.operation', 'Đầu năm')
    initial_number = fields.Many2one('financial.operation', 'Ghi sổ đầu năm')
    initial_preventive = fields.Many2one('financial.operation', 'Dự phòng đầu năm')
    last = fields.Many2one('financial.operation', 'Cuối năm')
    last_number = fields.Many2one('financial.operation', 'Ghi sổ cuối năm')
    last_preventive = fields.Many2one('financial.operation', 'Dự phòng cuối năm')

    # non_virtual_assets
    nc_vc_kt = fields.Many2one('financial.operation', 'Nhà cửa vật chất kiến trúc')
    equipment = fields.Many2one('financial.operation', 'Máy móc,thiết bị')
    pt_vt_td = fields.Many2one('financial.operation', 'Phương tiện vận tải,truyền dẫn')
    tb_cc_ql = fields.Many2one('financial.operation', 'Thiết bị công cụ quản lý')
    cc_sv = fields.Many2one('financial.operation', 'Cây cối,súc vật')
    kc_ht = fields.Many2one('financial.operation', 'Kết cấu,hạ tầng do NN ĐTXD')
    ts_khac = fields.Many2one('financial.operation', 'Tài sản cố định hữu hình khác')
    total = fields.Many2one('financial.operation', 'Tổng cộng')

    # virtual_assets
    q_s_d_d = fields.Many2one('financial.operation', 'Quyền sử dụng đất')
    q_ph = fields.Many2one('financial.operation', 'Quyền phát hành')
    copy_right = fields.Many2one('financial.operation', 'Bản quyền,bằng sáng chế')
    nh_hh = fields.Many2one('financial.operation', 'Nhẫn hiệu,hàng hóa')
    sort_ware = fields.Many2one('financial.operation', 'Phần mềm máy tính')
    license = fields.Many2one('financial.operation', 'Giấy phép')
    ts_vh = fields.Many2one('financial.operation', 'TSCĐ vô hình')

    # bds_invest
    up_in_year = fields.Many2one('financial.operation', 'Tăng trong năm')
    down_in_year = fields.Many2one('financial.operation', 'Tăng trong năm')

    # borrow
    lose_payment_last = fields.Many2one('financial.operation', 'Số không có khả năng trả nợ cuối năm')
    last_year_interest = fields.Many2one('financial.operation', 'Lãi cuối năm')
    lose_payment_init = fields.Many2one('financial.operation', 'Số không có khả năng trả nợ đầu năm')
    init_year_interest = fields.Many2one('financial.operation', 'Lãi đầu năm')

    # rest_debit
    rest_this_year = fields.Many2one('financial.operation', 'TT tiền thuê tài chính năm nay')
    pay_interest_this_year = fields.Many2one('financial.operation', 'Trả lãi thuê năm nay')
    pay_debit_this_year = fields.Many2one('financial.operation', 'Trả nợ gốc năm nay')
    rest_last_year = fields.Many2one('financial.operation', 'TT tiền thuê tài chính năm trước')
    pay_interest_last_year = fields.Many2one('financial.operation', 'Trả lãi thuê năm trước')
    pay_debit_last_year = fields.Many2one('financial.operation', 'Trả nợ gốc năm trước')

    # vat_other
    turn_in_year = fields.Many2one('financial.operation', 'Số phải nộp trong năm')
    real_turn_in_year = fields.Many2one('financial.operation', 'Số thực nộp trong năm')

    # csh
    von_gop_csh = fields.Many2one('financial.operation', 'Vốn góp chủ sở hữu')
    td_vcp = fields.Many2one('financial.operation', 'Thặng dư vốn cổ phần')
    cd_tp = fields.Many2one('financial.operation', 'Quyền chọn chuyển đổi trái phiếu')
    vk_csh = fields.Many2one('financial.operation', 'Vốn khác chủ sở hữu')
    cl_dg_ts = fields.Many2one('financial.operation', 'Chênh lệch đánh giá tài sản')
    cl_tg = fields.Many2one('financial.operation', 'Chênh lệch tỉ giá')
    lnst = fields.Many2one('financial.operation', 'LNST chưa phân phối và các quỹ')
    other_csh = fields.Many2one('financial.operation', 'Các khoản mục khác')

    # customer_pay
    last_year_can_pay = fields.Many2one('financial.operation', 'Có khả năng trả nợ cuối năm')
    initial_year_can_pay = fields.Many2one('financial.operation', 'Có khả năng trả nợ đầu năm')

    def query_execute(self, account_id=None, relate_account_id=None, date_from=None, date_to=None, type=None,
                      time=None):
        a = 0
        if date_from and date_to and account_id:
            #     date_from = date_from.strftime('%Y-%m-%d %H:%M:%S')
            #     date_to = date_to.strftime('%Y-%m-%d %H:%M:%S')
            if type == 'du_co_cuoi_ki' or type == 'du_no_cuoi_ki':
                # tong balance vn_account_move_line tk + tk con < date_to
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id 
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  aml.create_date <= %s
                                and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_to))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    if type == 'du_co_cuoi_ki':
                        return -result[0][0]
                    return result[0][0]
            elif type == 'du_co_cuoi_ki_chi_tiet'  or type == 'du_no_cuoi_ki_chi_tiet':
                # tong balance vn_account_move_line tk  < date_to
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id 
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  aml.create_date <= %s
                                and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code, date_to))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    if type == 'du_co_cuoi_ki_chi_tiet':
                        return -result[0][0]
                    return result[0][0]
            elif type == 'du_co_dau_ki' or type == 'du_no_dau_ki':
                # tong balance vn_account_move_line tk + con  < date_from
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  aml.create_date <= %s
                                and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_from))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    if type == 'du_co_dau_ki':
                        return -result[0][0]
                    return result[0][0]
            elif type == 'du_co_dau_ki_chi_tiet' or type == 'du_no_dau_ki_chi_tiet':
                # tong balance vn_account_move_line tk  < date_from
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  aml.create_date <= %s
                                and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code, date_from))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    if type == 'du_co_dau_ki_chi_tiet':
                        return -result[0][0]
                    return result[0][0]
            elif type == 'du_co_dau_ki_nt' or type == 'du_no_dau_ki_nt':
                # tong balance vn_account_move_line tk + con trong khoang   date_from - 1 nam -> date_from
                date_start = date_from - timedelta(days=365)
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s
                                and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%',date_start,date_from))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    if type == 'du_co_dau_ki_nt':
                        return -result[0][0]
                    return result[0][0]
            elif type == 'phat_sinh_co':
                # tong balance tk +tk con credit > 0 , trong khoang date_from -> date_to
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and  aml.create_date <= %s and aml.credit > 0
                                            and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_from,date_to))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    return -result[0][0]
            elif type == 'phat_sinh_co_nt':
                # tong balance tk +tk con credit > 0 , trong khoang date_from - 1 nam -> date_from
                date_start = date_from - timedelta(days=365)
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s and aml.credit > 0
                                            and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_start,date_from))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    return -result[0][0]
            elif type == 'phat_sinh_du':
                # tong balance tk +tk con co (relate_account_id hoac con cua relate_account_id) va trong khoang date_from -> date_to
                if relate_account_id:
                    query = """
                                SELECT sum(balance) as number
                                FROM vn_account_move_line aml
                                JOIN account_move move ON move.id = aml.move_id
                                WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s and aml.related_account_id in (SELECT id FROM account_account WHERE code like %s)
                                                                and move.state = 'posted'"""
                    self.env.cr.execute(query, (account_id.code + '%', date_from,date_to, relate_account_id.code + '%'))
                    result = self.env.cr.fetchall()
                    if result[0][0] is not None:
                        return result[0][0]
            elif type == 'phat_sinh_du_nt':
                # tong balance tk +tk con co (relate_account_id hoac con cua relate_account_id) va trong khoang date_from - 1 nam -> date_from
                date_start = date_from - timedelta(days=365)
                if relate_account_id:
                    query = """
                                   SELECT sum(balance) as number
                                   FROM vn_account_move_line aml
                                   JOIN account_move move ON move.id = aml.move_id
                                   WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s and aml.related_account_id in (SELECT id FROM account_account WHERE code like %s)
                                                                   and move.state = 'posted'"""
                    self.env.cr.execute(query, (account_id.code + '%', date_start, date_from, relate_account_id.code + '%'))
                    result = self.env.cr.fetchall()
                    if result[0][0] is not None:
                        return result[0][0]
            elif type == 'phat_sinh_no':
                # tong balance tk +tk con debit > 0 , trong khoang date_from -> date_to
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s and aml.debit > 0
                                            and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_from,date_to))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    return result[0][0]
            elif type == 'phat_sinh_no_nt':
                # tong balance tk +tk con debit > 0 , trong khoang date_from - 1 nam -> date_from
                date_start = date_from - timedelta(days=365)
                query = """
                            SELECT sum(balance) as number
                            FROM vn_account_move_line aml
                            JOIN account_move move ON move.id = aml.move_id
                            WHERE aml.account_id in (SELECT id FROM account_account WHERE code like %s) and  %s <= aml.create_date and aml.create_date <= %s and aml.debit > 0
                                            and move.state = 'posted'"""
                self.env.cr.execute(query, (account_id.code + '%', date_start,date_from))
                result = self.env.cr.fetchall()
                if result[0][0] is not None:
                    return result[0][0]

        return a

    def compute_line(self, date_from=None, date_to=None,time=None):
        result = []
        for rec in self:
            if rec.page == 'general':
                if rec.row and rec.col and rec.note:
                    result.append([rec.col + rec.row, rec.note])
                return result
            else:
                if rec.type == 'chi_tiet':
                    if rec.last:
                        value = 0
                        for line in rec.last.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if rec.last.row and rec.last.col:
                            result.append([rec.last.col + rec.last.row, value])
                    if rec.initial:
                        value = 0
                        for line in rec.initial.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if rec.initial.row and rec.initial.col:
                            result.append([rec.initial.col + rec.initial.row, value])
                    if rec.initial_number:
                        value = 0
                        for line in rec.initial_number.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if rec.initial_number.col and rec.initial_number.row:
                            result.append([rec.initial_number.col + rec.initial_number.row, value])
                    if rec.initial_preventive:
                        value = 0
                        for line in rec.initial_preventive.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if rec.initial_preventive.col and rec.initial_preventive.row:
                            result.append([rec.initial_preventive.col + rec.initial_preventive.row, value])
                    if rec.last_number:
                        value = 0
                        for line in rec.last_number.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.last_number.col + rec.last_number.row, value])
                    if rec.last_preventive:
                        value = 0
                        for line in rec.last_preventive.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.last_preventive.col + rec.last_preventive.row, value])
                    if rec.nc_vc_kt:
                        value = 0
                        for line in rec.nc_vc_kt.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.nc_vc_kt.col + rec.nc_vc_kt.row, value])
                    if rec.equipment:
                        value = 0
                        for line in rec.equipment.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.equipment.col + rec.equipment.row, value])
                    if rec.pt_vt_td:
                        value = 0
                        for line in rec.pt_vt_td.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.pt_vt_td.col + rec.pt_vt_td.row, value])
                    if rec.tb_cc_ql:
                        value = 0
                        for line in rec.tb_cc_ql.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.tb_cc_ql.col + rec.tb_cc_ql.row, value])
                    if rec.cc_sv:
                        value = 0
                        for line in rec.cc_sv.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.cc_sv.col + rec.cc_sv.row, value])
                    if rec.kc_ht:
                        value = 0
                        for line in rec.kc_ht.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.kc_ht.col + rec.kc_ht.row, value])
                    if rec.ts_khac:
                        value = 0
                        for line in rec.ts_khac.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.ts_khac.col + rec.ts_khac.row, value])
                    if rec.total:
                        value = 0
                        for line in rec.total.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.total.col + rec.total.row, value])
                    if rec.q_s_d_d:
                        value = 0
                        for line in rec.q_s_d_d.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.q_s_d_d.col + rec.q_s_d_d.row, value])
                    if rec.q_ph:
                        value = 0
                        for line in rec.q_ph.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.q_ph.col + rec.q_ph.row, value])
                    if rec.copy_right:
                        value = 0
                        for line in rec.copy_right.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.copy_right.col + rec.copy_right.row, value])
                    if rec.nh_hh:
                        value = 0
                        for line in rec.nh_hh.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.nh_hh.col + rec.nh_hh.row, value])
                    if rec.sort_ware:
                        value = 0
                        for line in rec.sort_ware.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.sort_ware.col + rec.sort_ware.row, value])
                    if rec.license:
                        value = 0
                        for line in rec.license.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.license.col + rec.license.row, value])
                    if rec.ts_vh:
                        value = 0
                        for line in rec.ts_vh.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.ts_vh.col + rec.ts_vh.row, value])
                    if rec.up_in_year:
                        value = 0
                        for line in rec.up_in_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.up_in_year.col + rec.up_in_year.row, value])
                    if rec.down_in_year:
                        value = 0
                        for line in rec.down_in_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.down_in_year.col + rec.down_in_year.row, value])
                    if rec.lose_payment_last:
                        value = 0
                        for line in rec.lose_payment_last.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.lose_payment_last.col + rec.lose_payment_last.row, value])
                    if rec.last_year_interest:
                        value = 0
                        for line in rec.last_year_interest.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.last_year_interest.col + rec.last_year_interest.row, value])
                    if rec.lose_payment_init:
                        value = 0
                        for line in rec.lose_payment_init.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.lose_payment_init.col + rec.lose_payment_init.row, value])
                    if rec.init_year_interest:
                        value = 0
                        for line in rec.init_year_interest.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.init_year_interest.col + rec.init_year_interest.row, value])
                    if rec.rest_this_year:
                        value = 0
                        for line in rec.rest_this_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.rest_this_year.col + rec.rest_this_year.row, value])
                    if rec.pay_interest_this_year:
                        value = 0
                        for line in rec.pay_interest_this_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.pay_interest_this_year.col + rec.pay_interest_this_year.row, value])
                    if rec.pay_debit_this_year:
                        value = 0
                        for line in rec.pay_debit_this_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.pay_debit_this_year.col + rec.pay_debit_this_year.row, value])
                    if rec.rest_last_year:
                        value = 0
                        for line in rec.rest_last_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.rest_last_year.col + rec.rest_last_year.row, value])
                    if rec.pay_interest_last_year:
                        value = 0
                        for line in rec.pay_interest_last_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.pay_interest_last_year.col + rec.pay_interest_last_year.row, value])
                    if rec.pay_debit_last_year:
                        value = 0
                        for line in rec.pay_debit_last_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.pay_debit_last_year.col + rec.pay_debit_last_year.row, value])
                    if rec.turn_in_year:
                        value = 0
                        for line in rec.turn_in_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.turn_in_year.col + rec.turn_in_year.row, value])
                    if rec.real_turn_in_year:
                        value = 0
                        for line in rec.real_turn_in_year.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.real_turn_in_year.col + rec.real_turn_in_year.row, value])
                    if rec.von_gop_csh:
                        value = 0
                        for line in rec.von_gop_csh.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.von_gop_csh.col + rec.von_gop_csh.row, value])
                    if rec.td_vcp:
                        value = 0
                        for line in rec.td_vcp.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.td_vcp.col + rec.td_vcp.row, value])
                    if rec.cd_tp:
                        value = 0
                        for line in rec.cd_tp.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.cd_tp.col + rec.cd_tp.row, value])
                    if rec.vk_csh:
                        value = 0
                        for line in rec.vk_csh.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.vk_csh.col + rec.vk_csh.row, value])
                    if rec.cl_dg_ts:
                        value = 0
                        for line in rec.cl_dg_ts.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.cl_dg_ts.col + rec.cl_dg_ts.row, value])
                    if rec.cl_tg:
                        value = 0
                        for line in rec.cl_tg.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.cl_tg.col + rec.cl_tg.row, value])
                    if rec.lnst:
                        value = 0
                        for line in rec.lnst.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.lnst.col + rec.lnst.row, value])
                    if rec.other_csh:
                        value = 0
                        for line in rec.other_csh.operation_lines:
                            if line.opera == 'plus':
                                value = value + rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                            else:
                                value = value - rec.query_execute(account_id=line.account_id,
                                                                  relate_account_id=line.account_id_balance,
                                                                  date_from=date_from, date_to=date_to, type=line.type,
                                                                  time=time)
                        if value:
                            result.append([rec.other_csh.col + rec.other_csh.row, value])
                if rec.type == 'tong_hop':
                    if rec.last:
                        value = 0
                        for line in rec.last.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from, date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.last.col + rec.last.row, value])
                    if rec.initial:
                        value = 0
                        for line in rec.initial.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from, date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.initial.col + rec.initial.row, value])
                    if rec.initial_preventive:
                        value = 0
                        for line in rec.initial_preventive.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.initial_preventive.col + rec.initial_preventive.row, value])
                    if rec.last_number:
                        value = 0
                        for line in rec.last_number.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.last_number.col + rec.last_number.row, value])
                    if rec.last_preventive:
                        value = 0
                        for line in rec.last_preventive.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.last_preventive.col + rec.last_preventive.row, value])
                    if rec.nc_vc_kt:
                        value = 0
                        for line in rec.nc_vc_kt.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.nc_vc_kt.col + rec.nc_vc_kt.row, value])
                    if rec.equipment:
                        value = 0
                        for line in rec.equipment.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.equipment.col + rec.equipment.row, value])
                    if rec.pt_vt_td:
                        value = 0
                        for line in rec.pt_vt_td.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.pt_vt_td.col + rec.pt_vt_td.row, value])
                    if rec.tb_cc_ql:
                        value = 0
                        for line in rec.tb_cc_ql.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.tb_cc_ql.col + rec.tb_cc_ql.row, value])
                    if rec.cc_sv:
                        value = 0
                        for line in rec.cc_sv.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.cc_sv.col + rec.cc_sv.row, value])
                    if rec.kc_ht:
                        value = 0
                        for line in rec.kc_ht.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.kc_ht.col + rec.kc_ht.row, value])
                    if rec.ts_khac:
                        value = 0
                        for line in rec.ts_khac.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.ts_khac.col + rec.ts_khac.row, value])
                    if rec.total:
                        value = 0
                        for line in rec.total.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.total.col + rec.total.row, value])
                    if rec.q_s_d_d:
                        value = 0
                        for line in rec.q_s_d_d.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.q_s_d_d.col + rec.q_s_d_d.row, value])
                    if rec.q_ph:
                        value = 0
                        for line in rec.q_ph.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.q_ph.col + rec.q_ph.row, value])
                    if rec.copy_right:
                        value = 0
                        for line in rec.copy_right.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.copy_right.col + rec.copy_right.row, value])
                    if rec.nh_hh:
                        value = 0
                        for line in rec.nh_hh.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.nh_hh.col + rec.nh_hh.row, value])
                    if rec.sort_ware:
                        value = 0
                        for line in rec.sort_ware.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.sort_ware.col + rec.sort_ware.row, value])
                    if rec.license:
                        value = 0
                        for line in rec.license.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.license.col + rec.license.row, value])
                    if rec.ts_vh:
                        value = 0
                        for line in rec.ts_vh.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.ts_vh.col + rec.ts_vh.row, value])
                    if rec.up_in_year:
                        value = 0
                        for line in rec.up_in_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.up_in_year.col + rec.up_in_year.row, value])
                    if rec.down_in_year:
                        value = 0
                        for line in rec.down_in_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.down_in_year.col + rec.down_in_year.row, value])
                    if rec.lose_payment_last:
                        value = 0
                        for line in rec.lose_payment_last.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.lose_payment_last.col + rec.lose_payment_last.row, value])
                    if rec.last_year_interest:
                        value = 0
                        for line in rec.last_year_interest.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.last_year_interest.col + rec.last_year_interest.row, value])
                    if rec.lose_payment_init:
                        value = 0
                        for line in rec.lose_payment_init.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.lose_payment_init.col + rec.lose_payment_init.row, value])
                    if rec.init_year_interest:
                        value = 0
                        for line in rec.init_year_interest.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.init_year_interest.col + rec.init_year_interest.row, value])
                    if rec.rest_this_year:
                        value = 0
                        for line in rec.rest_this_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.rest_this_year.col + rec.rest_this_year.row, value])
                    if rec.pay_interest_this_year:
                        value = 0
                        for line in rec.pay_interest_this_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.pay_interest_this_year.col + rec.pay_interest_this_year.row, value])
                    if rec.pay_debit_this_year:
                        value = 0
                        for line in rec.pay_debit_this_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.pay_debit_this_year.col + rec.pay_debit_this_year.row, value])
                    if rec.rest_last_year:
                        value = 0
                        for line in rec.rest_last_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.rest_last_year.col + rec.rest_last_year.row, value])
                    if rec.pay_interest_last_year:
                        value = 0
                        for line in rec.pay_interest_last_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.pay_interest_last_year.col + rec.pay_interest_last_year.row, value])
                    if rec.pay_debit_last_year:
                        value = 0
                        for line in rec.pay_debit_last_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.pay_debit_last_year.col + rec.pay_debit_last_year.row, value])
                    if rec.turn_in_year:
                        value = 0
                        for line in rec.turn_in_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.turn_in_year.col + rec.turn_in_year.row, value])
                    if rec.real_turn_in_year:
                        value = 0
                        for line in rec.real_turn_in_year.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.real_turn_in_year.col + rec.real_turn_in_year.row, value])
                    if rec.von_gop_csh:
                        value = 0
                        for line in rec.von_gop_csh.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.von_gop_csh.col + rec.von_gop_csh.row, value])
                    if rec.td_vcp:
                        value = 0
                        for line in rec.td_vcp.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.td_vcp.col + rec.td_vcp.row, value])
                    if rec.cd_tp:
                        value = 0
                        for line in rec.cd_tp.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.cd_tp.col + rec.cd_tp.row, value])
                    if rec.vk_csh:
                        value = 0
                        for line in rec.vk_csh.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.vk_csh.col + rec.vk_csh.row, value])
                    if rec.cl_dg_ts:
                        value = 0
                        for line in rec.cl_dg_ts.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.cl_dg_ts.col + rec.cl_dg_ts.row, value])
                    if rec.cl_tg:
                        value = 0
                        for line in rec.cl_tg.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.cl_tg.col + rec.cl_tg.row, value])
                    if rec.lnst:
                        value = 0
                        for line in rec.lnst.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.lnst.col + rec.lnst.row, value])
                    if rec.other_csh:
                        value = 0
                        for line in rec.other_csh.operation_lines:
                            if line.type in ['chi_tieu']:
                                data = self.env['financial.config.line'].search(
                                    [('code', '=', line.code)]).compute_line(date_from=date_from,
                                                                             date_to=date_to)
                                if line.opera == 'plus':
                                    for a in data:
                                        value += a[1]
                                else:
                                    for a in data:
                                        value -= a[1]
                        if value:
                            result.append([rec.other_csh.col + rec.other_csh.row, value])
            return result


class FinancialOperation(models.Model):
    _name = 'financial.operation'

    _order = 'id asc'
    _rec_name = 'ie_key'

    name = fields.Char('Tên')
    col = fields.Char("Thuộc cột")
    row = fields.Char("Thuộc hàng")
    operation_lines = fields.One2many('financial.operation.line', 'operation_line_id', 'Xây dựng công thức')
    type = fields.Selection([('dau_ki', 'Đầu kì'), ('cuoi_ki', 'Cuối kì')], string="Thời gian xét")
    ie_key = fields.Char()


class FinancialOperationLine(models.Model):
    _name = 'financial.operation.line'

    name = fields.Char('Tên')
    opera = fields.Selection([('minus', '-'), ('plus', '+')], string="Phép tính")
    note = fields.Char('Diễn giải')
    type = fields.Selection([('du_co_cuoi_ki', 'DUCOCK'),
                             ('du_co_cuoi_ki_chi_tiet', 'DUCOCK_ChitietTheoTK'),
                             ('du_co_cuoi_ki_chi_tiet_dt', 'DUCOCK_ChitietTheoTKvaDoiTuong'),
                             ('du_co_dau_ki', 'DUCODK'),
                             ('du_co_dau_ki_chi_tiet', 'DUCODK_ChitietTheoTK'),
                             ('du_co_dau_ki_chi_tiet_dt', 'DUCODK_ChitietTheoTKvaDoiTuong'),
                             ('du_co_dau_ki_nt', 'DUCODK_Namtruoc'),
                             ('du_no_cuoi_ki', 'DUNOCK'),
                             ('du_no_cuoi_ki_chi_tiet', 'DUNOCK_ChitietTheoTK'),
                             ('du_no_cuoi_ki_chi_tiet_dt', 'DUNOCK_ChitietTheoTKvaDoiTuong'),
                             ('du_no_dau_ki', 'DUNODK'),
                             ('du_no_dau_ki_chi_tiet', 'DUNODK_ChitietTheoTK'),
                             ('du_no_dau_ki_chi_tiet_dt', 'DUNODK_ChitietTheoTKvaDoiTuong'),
                             ('du_no_dau_ki_nt', 'DUNODK_Namtruoc'),
                             ('phat_sinh_co', 'PhatsinhCO'),
                             ('phat_sinh_co_nt', 'PhatsinhCO_Namtruoc'),
                             ('phat_sinh_du', 'PhatsinhDU'),
                             ('phat_sinh_du_nt', 'PhatsinhDU_Namtruoc'),
                             ('phat_sinh_du_vat', 'PhatsinhDUGiamVAT'),
                             ('phat_sinh_no', 'PhatsinhNO'),
                             ('phat_sinh_no_nt', 'PhatsinhNO_Namtruoc'),
                             ('phat_sinh_no_vat', 'PhatsinhNOGiamVAT'),
                             ('chi_tieu', 'Chỉ tiêu'),
                             ], string="Kí hiệu")
    code = fields.Char("Mã chỉ tiêu")
    account_id = fields.Many2one('account.account', 'Tài khoản')
    account_id_balance = fields.Many2one('account.account', 'Tài khoản đói ứng')
    operation_line_id = fields.Many2one('financial.operation')


class FinancialReport(models.TransientModel):
    _name = 'financial.out.report'
    _description = 'Financial Speech'

    period = fields.Selection([('month', 'Theo tháng'),
                               ('quarter', 'Theo quý'),
                               ('year', 'Theo năm'),
                               ],default='year', string="Kì chọn")
    date_from = fields.Date(
        string='Từ ngày',
        required=False)
    date_to = fields.Date(
        string='Đến ngày',
        required=False)

    def confirm(self):
        data = {}
        date_from = None
        date_to = None
        if self.date_from and self.date_to:
            date_from = datetime.combine(self.date_from, time(0, 0))
            date_to = datetime.combine(self.date_to, time(0, 0))
        financial = self.env['financial.note'].sudo().search([], limit=1)
        if financial:
            financial_lines = self.env['financial.config.line'].sudo().search(
                [('financial_note_id', '=', financial.id)])
            if len(financial_lines) > 0:
                for line in financial_lines:
                    line_data = line.compute_line(date_from=date_from, date_to=date_to,time=self.period)
                    if len(line_data) > 0 and type(line_data[0][1]) != str  and int(line_data[0][1]) > 0:
                        print(line.name)
                        print(line_data)
                    for dt in line_data:
                        data.update({
                            dt[0]: dt[1]
                        })
        if data:
            try:
                wb = load_workbook(get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                       'Thuyết minh BCTC theo thông tư 200.xlsx'))
                ws = wb['Sheet1']
                content = BytesIO()
                for key in data:
                    if ws[key].value is not None and str(ws[key].value) != '0':
                        ws[key] = ws[key].value + data[key]
                    else:
                        ws[key] = data[key]
                        ws[key].font = Font(name='Times New Roman', size=10)
                        ws[key].alignment = Alignment(horizontal='center')
                wb.save(content)
                out = base64.encodestring(content.getvalue())
                self.env['result.report.financial'].sudo().search([]).unlink()
                file_output = self.env['result.report.financial'].sudo().create({
                    'file': out,
                    'file_name': "Thuyết minh BCTC theo thông tư 200" + ".xlsx"
                })
                content.close()
                return {
                    'type': 'ir.actions.act_url',
                    'target': 'new',
                    'url': 'web/content/?model=' + file_output._name + '&id=' + str(
                        file_output.id) + '&field=file&download=true&filename=' + str(file_output.file_name),
                }
            except Exception as e:
                print(e)


class ResultReportFinancial(models.TransientModel):
    _name = "result.report.financial"

    file = fields.Binary("Result")
    file_name = fields.Char(string='Name')
