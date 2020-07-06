import base64
import logging
from copy import copy
from io import BytesIO
from datetime import datetime

from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class ExportWarehouseTag(models.Model):
    _name = 'export.warehouse.tag'

    date_from = fields.Date('Từ ngày')
    date_to = fields.Date('Đến ngày')
    warehouse_id = fields.Many2one('stock.location', string='Nhà kho')
    product_ids = fields.Many2many('product.product', string='Sản phẩm')

    @api.constrains('date_to', 'date_from')
    def _check_date_to_from_warehouse(self):
        for rec in self:
            if rec.date_to < rec.date_from:
                raise ValidationError(_('Ngày bắt đầu phải trước ngày kết thúc.'))

    def confirm_export_date(self):
        wb = load_workbook(get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                               'Thẻ kho.xlsx'))
        ws = wb['Sheet1']
        index = 1
        if len(self.product_ids.ids) > 1:
            index_sheet = 2
            so_sheet_them = len(self.product_ids.ids) - 1
            for i in range(so_sheet_them):
                new_sheet = wb.copy_worksheet(ws)
                new_sheet.title = 'Sheet' + str(index_sheet)
                index_sheet += 1
        for product in self.product_ids:
            self.fill_data(workbook=wb, worksheet=wb['Sheet' + str(index)], product=product)
            index += 1
        content = BytesIO()
        wb.save(content)
        out = base64.encodestring(content.getvalue())
        view = self.env.ref('advanced_vn_report.result_export_warehouse_tag_form_view')
        self.env['result.export.warehouse.tag'].sudo().search([]).unlink()
        file_output = self.env['result.export.warehouse.tag'].sudo().create({
            'file': out,
            'file_name': "Thẻ kho.xlsx"
        })
        content.close()
        return {
            'name': "Thẻ kho",
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'result.export.warehouse.tag',
            'target': 'new',
            'view_id': view.id,
            'res_id': file_output.id
        }

    def fill_data(self, workbook, worksheet, product):
        worksheet['A5'] = "Ngày lập thẻ: " + datetime.today().strftime("%d/%m/%Y")
        worksheet['E7'] = product.name
        worksheet['E8'] = product.uom_id.name if product.uom_id else None
        worksheet[
            'E9'] = product.product_tmpl_id.default_code if product.product_tmpl_id and product.product_tmpl_id.default_code else None
        picking_done = self.env['stock.picking'].sudo().search(
            [('state', '=', 'done'), ('date_done', '>=', self.date_from), ('date_done', '<=', self.date_to)])
        row = 13
        col = 1
        index = 1
        global_style = copy(worksheet['B14']._style)
        so_dong = 0
        ton_dau_ki = product._compute_quantities_dict(
            self._context.get('lot_id'),
            self._context.get('owner_id'),
            self._context.get('package_id'), self.date_from, self.date_to)[product.id]['qty_available']
        if picking_done:
            for rec in picking_done:
                if rec.picking_type_id.code == 'incoming':
                    if rec.location_dest_id.id == self.warehouse_id.id:
                        if rec.move_ids_without_package:
                            for line in rec.move_ids_without_package:
                                if line.product_id.id == product.id:
                                    so_dong += 1
                elif rec.picking_type_id.code == 'outgoing':
                    if rec.location_id.id == self.warehouse_id.id:
                        if rec.move_line_ids_without_package:
                            for line in rec.move_ids_without_package:
                                if line.product_id.id == product.id:
                                    so_dong += 1
            # print(so_dong)
            if so_dong >= 5:
                so_dong_them = so_dong - 4
                start_insert_row = 17
                while so_dong_them > 0:
                    worksheet.insert_rows(start_insert_row)
                    for coltmp in range(10):
                        worksheet.cell(row=start_insert_row, column=coltmp + 1)._style = global_style
                    start_insert_row = start_insert_row + 1
                    so_dong_them -= 1
            tong_nhap_cuoi_ki = 0
            tong_xuat_cuoi_ki = 0
            for picking in picking_done:
                if picking.picking_type_id.code == 'incoming':
                    if picking.location_dest_id.id == self.warehouse_id.id:
                        if picking.move_ids_without_package:
                            for line in picking.move_ids_without_package:
                                if line.product_id.id == product.id:
                                    worksheet.cell(row=row, column=col).value = index
                                    index += 1
                                    worksheet.cell(row=row, column=col + 1).value = picking.date_done.strftime(
                                        "%d/%m/%Y")
                                    worksheet.cell(row=row, column=col + 2).value = picking.name
                                    worksheet.cell(row=row,
                                                   column=col + 4).value = picking.purchase_id.name + " mua hàng từ " + picking.purchase_id.partner_id.name if picking.purchase_id else None
                                    worksheet.cell(row=row, column=col + 5).value = picking.date_done.strftime(
                                        "%d/%m/%Y")
                                    worksheet.cell(row=row, column=col + 6).value = line.quantity_done
                                    tong_nhap_cuoi_ki += line.quantity_done
                                    ton_dau_ki += line.quantity_done
                                    worksheet.cell(row=row, column=col + 8).value = ton_dau_ki
                                    row += 1
                                    col = 1
                elif picking.picking_type_id.code == 'outgoing':
                    if picking.location_id.id == self.warehouse_id.id:
                        if picking.move_line_ids_without_package:
                            for line in picking.move_line_ids_without_package:
                                if line.product_id.id == product.id:
                                    worksheet.cell(row=row, column=col).value = index
                                    index += 1
                                    worksheet.cell(row=row, column=col + 1).value = picking.date_done.strftime(
                                        "%d/%m/%Y")
                                    worksheet.cell(row=row, column=col + 3).value = picking.name
                                    worksheet.cell(row=row,
                                                   column=col + 4).value = picking.sale_id.name + " bán hàng cho " + picking.sale_id.partner_id.name if picking.purchase_id else None
                                    worksheet.cell(row=row, column=col + 5).value = picking.date_done.strftime(
                                        "%d/%m/%Y")
                                    worksheet.cell(row=row, column=col + 7).value = line.qty_done
                                    tong_xuat_cuoi_ki += line.qty_done
                                    ton_dau_ki -= line.qty_done
                                    worksheet.cell(row=row, column=col + 8).value = ton_dau_ki
                                    row += 1
                                    col = 1
            worksheet.cell(row=row, column=col + 4).value = 'Cộng cuối kỳ'
            worksheet.cell(row=row, column=col + 6).value = tong_nhap_cuoi_ki
            worksheet.cell(row=row, column=col + 7).value = tong_xuat_cuoi_ki
            worksheet.cell(row=row, column=col + 8).value = ton_dau_ki
